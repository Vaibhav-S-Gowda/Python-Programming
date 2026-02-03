import os
from openpyxl import Workbook , load_workbook

filename = "openpyxl_master.xlsx"

if not os.path.exists(filename):
    wb = Workbook()

    print(type(wb))

    ws = wb.active # Default sheet
    ws.title = "Students"

    wb.save(filename)
    print("Workbook created successfully")

# 3. Load an Existing Workbook
wb = load_workbook(filename)
print("Sheets: ", wb.sheetnames)

# 4. Writing Data into Cells
ws = wb["Students"]
if ws["A1"].value != "ID":

    ws["A1"] = "ID"
    ws["B1"] = "Name"
    ws["C1"] = "Marks"

    ws["A2"] = 101
    ws["B2"] = "Jhonny"
    ws["C2"] = 95

    wb.save(filename)
    print("Data written successfully")

# # 5. Append Rows Automatically
# existing_ids = [cell.value for cell in ws["A"]]

# if 102 not in existing_ids: 
#     ws.append([102, "Rahul", 88])
# if 103 not in existing_ids:
#     ws.append([103, "Anitha", 83])
# wb.save(filename)
# print("Rows appended successfully")

# 6. Reading Cell Data
'''print("A2 = ", ws["A2"].value)
print("A3 = ", ws["A3"].value)
print("A4 = ", ws["A4"].value)
print("B2 = ", ws["B2"].value)
print("B3 = ", ws["B3"].value)
print("B4 = ", ws["B4"].value)
print("C2 = ", ws["C2"].value)
print("C3 = ", ws["C3"].value)
print("C4 = ", ws["C4"].value)'''

# 7. Reading All Rows and Columns
for row in ws.iter_rows(values_only=True):
    print(row)

# 8. Get Row and Column Counts
print("Total rows: ", ws.max_row)
print("Total columns: ", ws.max_column)

# # 9. Create Multiple Sheets
# if "Teachers" not in wb.sheetnames:
#     wb.create_sheet("Teachers")
# if "Courses" not in wb.sheetnames:
#     wb.create_sheet("Courses")

# wb.save(filename)
# print("Sheets added: ",wb.sheetnames)

# 10. Delete a Sheet
for sheet in wb.sheetnames:
    if sheet not in ["Students", "Faculty"]:
        del wb[sheet]

    wb.save(filename)
    print("Sheet deleted")

# 11. Rename a Sheet
if "Teachers" in wb.sheetnames:
    ws = wb["Teachers"]
    ws.title = "Faculty"
    wb.save(filename)
    print("Sheet renamed")

# # 12. Formatting Cells (Font, Color, Alignment)
# from openpyxl.styles import Font, Alignment
# for cell in ["A1", "B1", "C1"]:
#     ws[cell].font = Font(bold=True, color="FF0000")
#     ws[cell].alignment = Alignment(horizontal="center")
# wb.save(filename)
# print("Formatting applied")


# 13. Fill Background Color
from openpyxl.styles import PatternFill

wb = load_workbook("openpyxl_master.xlsx")
ws = wb["Students"]

fill = PatternFill(start_color="FFFF00", fill_type="solid")

ws["B1"].fill = fill

wb.save("openpyxl_master.xlsx")
print("Cell fill added")



# 14. Adjust Column Width
ws.column_dimensions["B"].width = 20

wb.save("openpyxl_master.xlsx")
print("Column width adjusted")


# 15. Merge and Unmerge Cells
ws.merge_cells("E1:G1")
ws["E1"] = "Merged Title"

wb.save("openpyxl_master.xlsx")
print("Cells merged")



# 17. Add Filters
ws.auto_filter.ref = "A1:D4"

wb.save("openpyxl_master.xlsx")
print("Filter added")


# 18. Freeze Panes
ws.freeze_panes = "A2"

wb.save("openpyxl_master.xlsx")
print("Panes frozen")




# 20. Add Charts
from openpyxl.chart import BarChart, Reference

if len(ws._charts) == 0:

    chart = BarChart()
    chart.title = "Marks Chart"

    # Select all rows and columns with data
    data = Reference(
        ws,
        min_col=1,
        min_row=1,
        max_col=ws.max_column,
        max_row=ws.max_row
    )

    chart.add_data(data, titles_from_data=True)

    ws.add_chart(chart, "F5")
    wb.save("openpyxl_master.xlsx")

    print("Chart added successfully with all rows and columns")

else:
    print("Chart already exists, skipping...")




# 22. Protect Sheet
ws.protection.sheet = True
ws.protection.set_password("1234")

wb.save("openpyxl_master.xlsx")
print("Sheet protected successfully")
